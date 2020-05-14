import face_recognition
import cv2
from imutils import paths
import os
from datetime import date
import openpyxl
from collections import defaultdict
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font



font = Font(bold=True)
font1=Font(italic=True,color='FF595959')

sheet_add="C://Users//P.Harish Kumar//Desktop//Project Attendence//xlsheet"
wb = openpyxl.load_workbook(sheet_add+"//Attendance.xlsx")
s1 = wb.get_sheet_by_name('Sheet1')

b=defaultdict(lambda:0)

m=2

while(s1.cell(m,1).value):
    print(s1.cell(m,1).value)
    m=m+1


today=date.today()
temp1=str(today.day)+"/"+str(today.month)+"/"+str(today.year)

m=2
while(s1.cell(1,m).value):
    if(s1.cell(1,m).value==temp1): break
    m=m+1

if(s1.cell(1,m).value==None):
    s1.cell(1,m).value=temp1
    s1.cell(1,m).font=font
else:
	k=2
	temp=s1.cell(k,m).value
	while(temp):
		if(temp=="P"): b[s1.cell(k,1).value]+=10
		k+=1
		temp=s1.cell(k,m).value

cc=0


imagePaths = list(paths.list_images("C://Users//P.Harish Kumar//Desktop//Project Attendence//Pictures"))

known_face_encodings = []
known_face_names = []
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True

for (i, imagePath) in enumerate(imagePaths):
	print("[INFO] processing image {}/{}".format(i + 1,
		len(imagePaths)))
	name = imagePath.split(os.path.sep)[-2]


	image = cv2.imread(imagePath)
	rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)


	boxes = face_recognition.face_locations(rgb,2)


	encodings = face_recognition.face_encodings(rgb, boxes)


	for encoding in encodings:

		known_face_encodings.append(encoding)
		known_face_names.append(name)


video_capture = cv2.VideoCapture(0)

while True:
    ret, frame = video_capture.read()

    cc=cc+1
    
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

    rgb_small_frame = small_frame[:, :, ::-1]


    if process_this_frame:
        face_locations = face_recognition.face_locations(rgb_small_frame,2)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        face_names = []
        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"

            if True in matches:
                first_match_index = matches.index(True)
                name = known_face_names[first_match_index]
                b[name]+=1
            face_names.append(name)

    process_this_frame = not process_this_frame


    # Display the results
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4

        if(b[name]<5): 
            rec=(255,0,0)
            name=name+" (Abs)"	
        else: 
            rec=(0,255,0)
            name=name+" (Prs)"

        cv2.rectangle(frame, (left-25, top-25), (right+25, bottom+25), rec, 2)

        cv2.rectangle(frame, (left-25, bottom), (right+25, bottom+25), rec, cv2.FILLED)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name.split(".")[1], (left -21, bottom +19), font, 0.7, (255, 255, 255), 1)

    cv2.imshow('Video', frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

video_capture.release()
cv2.destroyAllWindows()

j=2
while(1):
    if(s1.cell(j,1).value==None): break

    val=b[s1.cell(j,1).value]
    if(val>4):
        k="P"
    else:
        k="A"
    if(s1.cell(j,m).value!="P"):
   	    s1.cell(j,m).value=k
    j=j+1

wb.save(sheet_add+"//Attendance.xlsx")