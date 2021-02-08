from flask import Flask, render_template, request,redirect,send_file
import matplotlib.pyplot as plt
import os,shutil
import numpy as np
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
import forms
import face_recognition
import cv2
from imutils import paths
import os
from datetime import date
import openpyxl
from collections import defaultdict
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


def start1():

	global known_face_encodings,known_face_names,face_locations,face_encodings,face_names,process_this_frame,sheet_add,font2,font1,ad_key

	font1 = Font(bold=True)
	font2=Font(italic=True,color='FF595959')


	sheet_add="E://Projects//Project Attendence//xlsheet//"
	wb = openpyxl.load_workbook("E://Projects//Project Attendence//ad_key.xlsx")
	s1 = wb.get_sheet_by_name('Sheet1')
	ad_key=defaultdict(lambda: "0")

	m=1

	while(s1.cell(m,1).value):
	    ad_key[s1.cell(m,2).value]=s1.cell(m,1).value
	    m=m+1


	#Preparation of encodings

	imagePaths = list(paths.list_images("E://Projects//Project Attendence//Pictures"))

	known_face_encodings = []
	known_face_names = []
	face_locations = []
	face_encodings = []
	face_names = []
	process_this_frame = True

	for (i, imagePath) in enumerate(imagePaths):

		print("[INFO] processing image {}/{}".format(i + 1,
			len(imagePaths)))
		name = imagePath.split(os.path.sep)[-2]


		image = cv2.imread(imagePath)
		rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)


		boxes = face_recognition.face_locations(rgb,2)


		encodings = face_recognition.face_encodings(rgb, boxes)


		for encoding in encodings:

			known_face_encodings.append(encoding)
			known_face_names.append(name)


#frame analysation for facial recognition

def face_rec(subj):

	global known_face_encodings,known_face_names,face_locations,face_encodings,face_names,process_this_frame,sheet_add,font2,font1

	wb = openpyxl.load_workbook(sheet_add+subj+".xlsx")
	s1 = wb.get_sheet_by_name('Sheet1')

	b=defaultdict(lambda:0)

	m=2

	while(s1.cell(m,1).value):

	    print(s1.cell(m,1).value)
	    m=m+1


	today=date.today()
	temp1=str(today.day)+"/"+str(today.month)+"/"+str(today.year)

	m=2

	while(s1.cell(1,m).value):

	    if(s1.cell(1,m).value==temp1): break
	    m=m+1

	if(s1.cell(1,m).value==None):

	    s1.cell(1,m).value=temp1
	    s1.cell(1,m).font=font1

	else:

		k=2
		temp=s1.cell(k,m).value

		while(temp):

			if(temp=="P"): b[s1.cell(k,1).value]+=10
			k+=1
			temp=s1.cell(k,m).value

	video_capture = cv2.VideoCapture(0)

	cc=0

	while True:
	    ret, frame = video_capture.read()

	    cc=cc+1
	    
	    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

	    rgb_small_frame = small_frame[:, :, ::-1]


	    if process_this_frame:

	        face_locations = face_recognition.face_locations(rgb_small_frame,2)
	        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

	        face_names = []

	        for face_encoding in face_encodings:

	            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
	            name = "Unknown"

	            if True in matches:

	                first_match_index = matches.index(True)
	                name = known_face_names[first_match_index]
	                b[name]+=1

	            face_names.append(name)

	    process_this_frame = not process_this_frame


	    # Display the results
	    font = cv2.FONT_HERSHEY_DUPLEX

	    for (top, right, bottom, left), name in zip(face_locations, face_names):

	        top *= 4
	        right *= 4
	        bottom *= 4
	        left *= 4

	        if(b[name]<5): 

	            rec=(255,0,0)
	            name=name+" (Abs)"	

	        else: 

	            rec=(0,255,0)
	            name=name+" (Prs)"

	        cv2.rectangle(frame, (left-25, top-25), (right+25, bottom+25), rec, 2)

	        cv2.rectangle(frame, (left-25, bottom), (right+25, bottom+25), rec, cv2.FILLED)
	        cv2.putText(frame, name.split(".")[1], (left -21, bottom +19), font, 0.7, (255, 255, 255), 1)

	    cv2.putText(frame,"Press Q to exit", (5, 18), font, 1.2, (255, 255, 255), 1)
	    cv2.imshow('Video', frame)

	    if cv2.waitKey(1) & 0xFF == ord('q'):
	        break

	video_capture.release()
	cv2.destroyAllWindows()

	j=2

	while(1):

	    if(s1.cell(j,1).value==None): break

	    val=b[s1.cell(j,1).value]

	    if(val>4):

	        k="P"

	    else:

	        k="A"

	    if(s1.cell(j,m).value!="P"):

	   	    s1.cell(j,m).value=k

	    j=j+1

	wb.save(sheet_add+subj+".xlsx")

def new_u(subj):

	global known_face_names,font2,font1

	wb = openpyxl.Workbook() 
	sheet = wb.active 
	sheet.title = "Sheet1"

	s1 = wb.get_sheet_by_name('Sheet1')
	s1.cell(1,1).value="NM/DY"
	s1.cell(1,1).font=font1

	for i in range(len(known_face_names)):

		s1.cell(i+2,1).value=known_face_names[i]
		s1.cell(i+2,1).font=font2

	wb.save(sheet_add+subj+".xlsx")

def view_att(dt):

	global sheet_add,lgn
	dt=dt.split("-")
	dt=str(int(dt[2]))+"/"+str(int(dt[1]))+"/"+dt[0]
	print(dt)

	pres=[]
	abse=[]
	wb = openpyxl.load_workbook(sheet_add+lgn.subject+".xlsx")
	s1 = wb.get_sheet_by_name('Sheet1')
	m=2
    
	t=s1.cell(1,m).value

	while(t):

	    if(t==dt): break
	    m=m+1
	    t=s1.cell(1,m).value

	if t!=dt: return
    
	j=2
	t=s1.cell(j,m).value

	while(t):

	    if(t=="P"): pres.append(s1.cell(j,1).value.split(".")[0])
	    else: abse.append(s1.cell(j,1).value.split(".")[0])
	    j=j+1
	    t=s1.cell(j,m).value

	pres=",".join(pres)
	abse=",".join(abse)

	if len(pres)==0: pres="None"
	if len(abse)==0: abse="None"

	wb.save(sheet_add+lgn.subject+".xlsx")

	return [pres,abse]

def edit_att(dt,pres1,abse1):

	global sheet_add,lgn

	dt=dt.split("-")
	dt=str(int(dt[2]))+"/"+str(int(dt[1]))+"/"+dt[0]

	pres1=pres1.split(",")
	abse1=abse1.split(",")

	pres=[]
	abse=[]

	wb = openpyxl.load_workbook(sheet_add+lgn.subject+".xlsx")
	s1 = wb.get_sheet_by_name('Sheet1')

	m=2
    
	t=s1.cell(1,m).value

	while(t):

	    if(t==dt): break
	    m=m+1
	    t=s1.cell(1,m).value

	if t!=dt: return
    
	j=2

	t=s1.cell(j,m).value

	while(t):

		temp=s1.cell(j,1).value.split(".")[0]

		if(temp in pres1): 

			s1.cell(j,m).value="P"
			t="P"

		if(temp in abse1): 

			s1.cell(j,m).value="A"
			t="A"

		if(t=="P"): pres.append(temp)

		else: abse.append(temp)

		j=j+1

		t=s1.cell(j,m).value

	pres=",".join(pres)
	abse=",".join(abse)

	if len(pres)==0: pres="None"
	if len(abse)==0: abse="None"

	wb.save(sheet_add+lgn.subject+".xlsx")

	return [pres,abse]


# Flask construction

app = Flask(__name__)

app.config['SECRET_KEY'] = 'mysecret'

basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///'+os.path.join(basedir, 'data.sqlite')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False


#Preparing teacher's database
db = SQLAlchemy(app)
Migrate(app,db)

class Att_data(db.Model):

    __tablename__ = 'Att_data'

    user = db.Column(db.String(30), primary_key=True)
    password= db.Column(db.String(30))
    email = db.Column(db.String(64))
    subject=db.Column(db.String(64))
    key=db.Column(db.String(5))

    def __init__(self, email, user, password,key,subject):

        self.email = email
        self.user = user
        self.password = password
        self.key=key
        self.subject=subject

    def __repr__(self):

        return f"Username {self.user} , password {self.password}"

db.create_all()


@app.route('/',methods=['GET','POST'])
def login():

    global lgn,ad_key

    form1=forms.login1()
    form2=forms.register1()

    if form1.submit.data==True:

        user=Att_data.query.get(form1.user1.data)

        if user!=None and user.password==form1.pass1.data:

            lgn=user
            check=0
            return render_template('button.html',lgn=lgn)

        else:

            return render_template('login.html',check=3)
    
    if form2.submit1.data==True:

        user=Att_data.query.get(form2.user.data)

        if user==None and form2.pass2.data==form2.pass3.data and form2.key.data==ad_key[form2.subject.data]:

            new_user=Att_data(form2.email.data,form2.user.data,form2.pass2.data,form2.key.data,form2.subject.data)
            db.session.add(new_user)
            db.session.commit()
            temp=form2.subject.data
            new_u(temp)
            return render_template('login.html',check=1)

        else:

            return render_template('login.html',check=2)
        

    return render_template('login.html',check=0)

@app.route('/button',methods=['GET','POST'])
def button():

    global lgn,check

    return render_template('button.html',lgn=lgn)


@app.route('/cap_att')
def cap_att():

    global lgn,check

    face_rec(lgn.subject)
    return render_template('button.html',lgn=lgn)

@app.route('/view_form',methods=['GET','POST'])
def view_form():

    global lgn,check

    form1=forms.view_f()
    att=None
    check=0

    if form1.submit2.data==True:

    	att=view_att(form1.date.data)
    	check=1

    return render_template('view_form.html',lgn=lgn,att=att,check=check)

@app.route('/edit_form',methods=['GET','POST'])
def edit_form():

    global lgn,check

    form1=forms.edit_f()
    att=None
    check=0

    if form1.submit3.data==True:

    	if form1.epass.data==lgn.password and form1.ekey.data==lgn.key:

    		print(form1.epres.data,form1.eabse.data)
    		att=edit_att(form1.date1.data,form1.epres.data,form1.eabse.data)

    	check=1

    return render_template('edit_form.html',lgn=lgn,att=att,check=check)

@app.route('/download')
def download():

    global lgn,check

    return send_file(sheet_add+lgn.subject+".xlsx", as_attachment=True)


@app.route('/about')
def about():

    return render_template('about.html')



@app.route('/contact')
def contact():

    return render_template('contact.html')


@app.errorhandler(404)
def page_not_found(e):

    return render_template('sorry.html',check=0), 404


if __name__ == '__main__':
	
	start1()
	app.run(debug=True)